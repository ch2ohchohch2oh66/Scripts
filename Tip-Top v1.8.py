import datetime
import logging
import re
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import time

# 定义一个字典来存储全局变量
global_vars = {}


def set_global(k, v):
    global_vars[k] = v


def get_global(k):
    return global_vars.get(k, None)  # 如果键不存在，返回None


# 配置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
# 创建日志记录器
logger = logging.getLogger(__name__)


# 定义一个函数，用于从top文件中提取符合条件的数据
def extract_top_data(pid_list, command_list):
    logger.info(f'开始处理top数据文件：{top_file_path}')
    # 打开txt文件并逐行读取数据
    with open(top_file_path, 'r') as file:
        # 初始化字典，用于存储CPU和MEM信息
        cpu_data = {'time': [], 'us': [], 'sy': [], 'ni': [], 'id': [], 'wa': [], 'hi': [], 'si': [], 'st': [], 'used': []}
        mem_data = {'time': [], 'total': [], 'free': [], 'used': [], 'buff/cache': [], 'nonfree rate': [], 'used rate': []}
        free_index = 1
        used_index = 2
        isInit = True
        current_time = None  # 用于存储当前时间

        # 对于一体化安装环境，要提取被测服务对应的PID的CPU和MEM数据
        pid_data = {}  # 存储PID对应的CPU和MEM数据
        summary_pid_cpu_data = {}  # 汇总被测服务对应的PID的CPU数据
        summary_pid_mem_data = {}  # 汇总被测服务对应的PID的MEM数据

        # 初始化process_list
        process_list = []
        if pid_list:
            process_list = pid_list
            logger.info(f'待提取PID: {process_list}')
        elif command_list:
            process_list = command_list
            logger.info(f'待提取COMMAND: {process_list}')
        else:
            logger.info('no process be selected!')

        if process_list:
            for process in process_list:
                sheet_name = f'PROCESS_{process}'  # pid作为worksheet或column的名称
                if sheet_name not in pid_data:
                    pid_data[sheet_name] = {'%CPU': [], '%MEM': []}
                if sheet_name not in summary_pid_cpu_data:
                    summary_pid_cpu_data[sheet_name] = []
                if sheet_name not in summary_pid_mem_data:
                    summary_pid_mem_data[sheet_name] = []

        for line in file:
            # 提取时间信息
            if line.startswith('top -'):
                current_time = line.split()[2]  # 获取时间部分
                logger.debug(f'当前时间: {current_time}')

            if line.startswith('%Cpu(s)') or line.startswith('Cpu(s)'):
                logger.debug(f'开始处理CPU信息：{line.strip()}')
                # 如果找到CPU参数，则将其添加到CPU参数字典中
                cpu_data_str = re.findall(r'\b\d+\.\d+\b', line)
                cpu_data_num = [float(num) for num in cpu_data_str]
                logger.debug(f'CPU信息匹配结果：{cpu_data_num}')

                # 计算CPU使用率100-idle
                used = 100 - cpu_data_num[3]
                cpu_data['time'].append(current_time)  # 使用提取的时间
                cpu_data['us'].append(cpu_data_num[0])
                cpu_data['sy'].append(cpu_data_num[1])
                cpu_data['ni'].append(cpu_data_num[2])
                cpu_data['id'].append(cpu_data_num[3])
                cpu_data['wa'].append(cpu_data_num[4])
                cpu_data['hi'].append(cpu_data_num[5])
                cpu_data['si'].append(cpu_data_num[6])
                cpu_data['st'].append(cpu_data_num[7])
                cpu_data['used'].append(used)

            elif line.startswith('MiB Mem') or line.startswith('KiB Mem') or line.startswith('Mem'):
                logger.debug(f'开始处理MEM信息：{line.strip()}')
                # 处理mem数据：兼容整数和小数两种情况
                mem_data_str = re.findall(r'\b\d+(?:\.\d+)?\b', line) or re.findall(r'\d+', line)
                mem_data_num = [float(num) for num in mem_data_str]
                logger.debug(f'MEM信息匹配结果：{mem_data_num}')

                # 确定free和used的索引
                if isInit:
                    original_free_index = line.index('free')
                    original_used_index = line.index('used')
                    if original_free_index > original_used_index:
                        logger.info('更新free/used位置')
                        free_index = 2
                        used_index = 1

                    isInit = False
                free = mem_data_num[free_index]
                used = mem_data_num[used_index]
                total = mem_data_num[0]
                buff_cache = mem_data_num[3]

                # 计算MEM使用率，nonfree rate：(total-free)/total
                nonfree_rate = (total - free) * 100 / total
                mem_data['time'].append(current_time)  # 使用提取的时间
                mem_data['total'].append(total)
                mem_data['free'].append(free)
                mem_data['used'].append(used)
                mem_data['buff/cache'].append(buff_cache)
                mem_data['nonfree rate'].append(round(nonfree_rate, 1))
                mem_data['used rate'].append(round(used * 100 / total, 1))
            elif process_list:
                # 匹配PID\CUP\MEM\TIME\COMMAND
                # match = re.search(rf'(\d+)\s+.*?(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+:\d+\.\d+)\s+(.+?)\n', line)
                match = re.search(rf'(\d+)\s+.*?(\d+\.\d+)\s+(\d+\.\d+)\s+(\d+:\d+(?:\.\d+)?)\s+(.+?)\n', line)

                # 并不是所有被测PID都会同时启动，且未加载有效业务前部分PID不断闪现，对于暂时消失的PID设为空，以让真正有效的数据保持对齐
                if not match and line.strip().startswith('PID') and line.strip().endswith('COMMAND'):
                    logger.debug('匹配到PID和COMMAND标题栏，开始初始化进程的datasheet')
                    for process in process_list:
                        sheet_name = f'PROCESS_{process}'
                        logger.debug(f'初始化datasheet: {sheet_name}')
                        # 修改为使用最近一次的值，而不是设置为0
                        last_cpu_value = pid_data[sheet_name]['%CPU'][-1] if pid_data[sheet_name]['%CPU'] else 0
                        last_mem_value = pid_data[sheet_name]['%MEM'][-1] if pid_data[sheet_name]['%MEM'] else 0
                        pid_data[sheet_name]['%CPU'].append(last_cpu_value)
                        pid_data[sheet_name]['%MEM'].append(last_mem_value)
                        summary_pid_cpu_data[sheet_name].append(last_cpu_value)
                        summary_pid_mem_data[sheet_name].append(last_mem_value)
                elif match:
                    pid, cpu, mem, time, command = match.groups()

                    if pid_list:
                        process = int(pid)
                    elif command_list:
                        process = command.strip()

                    if process in process_list:
                        # logger.info(f'匹配到进程：{process}, 匹配结果: {match.groups()}')
                        sheet_name = f'PROCESS_{process}'
                        if cpu:
                            pid_data[sheet_name]['%CPU'][-1] = float(cpu)
                            summary_pid_cpu_data[sheet_name][-1] = float(cpu)
                        if mem:
                            pid_data[sheet_name]['%MEM'][-1] = float(mem)
                            summary_pid_mem_data[sheet_name][-1] = float(mem)
    logger.info(f'完成处理top数据文件：{top_file_path}')

    if isAvg:
        # 对 cpu_data 和 mem_data 添加平均值
        for key in cpu_data.keys():
            if key == 'time':  # 不计算时间的平均值
                cpu_data[key].insert(0, 'Avg-->')
            else:
                cpu_data[key].insert(0, calculate_average(cpu_data[key], startIndex, endIndex))

        for key in mem_data.keys():
            if key == 'time':  # 不计算时间的平均值
                mem_data[key].insert(0, 'Avg-->')
            else:
                mem_data[key].insert(0, calculate_average(mem_data[key], startIndex, endIndex))

        # 求取summary_pid_cpu_data和summary_pid_mem_data平均值，并插入到每个数据的顶部
        for process in process_list:
            summary_pid_data_header = f'PROCESS_{process}'
            summary_pid_cpu = summary_pid_cpu_data[summary_pid_data_header]
            summary_pid_mem = summary_pid_mem_data[summary_pid_data_header]

            # 计算平均值
            avg_summary_pid_cpu = calculate_average(summary_pid_cpu, startIndex, endIndex)
            avg_summary_pid_mem = calculate_average(summary_pid_mem, startIndex, endIndex)

            # 将平均值插入到顶部
            summary_pid_cpu_data[summary_pid_data_header].insert(0, avg_summary_pid_cpu)
            summary_pid_mem_data[summary_pid_data_header].insert(0, avg_summary_pid_mem)

    # 将处理后的数据写入全局变量
    set_global('cpu_data', cpu_data)
    set_global('mem_data', mem_data)
    set_global('pid_data', pid_data)
    set_global('summary_pid_cpu_data', summary_pid_cpu_data)
    set_global('summary_pid_mem_data', summary_pid_mem_data)


# 从sar文件中提取数据，该方法适用于过滤了单个网卡或网口的数据，如：nohup sar -n DEV 1  | grep 'enp1s0' >> sar.txt &
def extract_sar_data():
    logger.info(f'开始处理sar数据文件：{sar_file_path}')
    # 打开txt文件并逐行读取数据
    with open(sar_file_path, 'r', encoding='utf-8') as file:
        # 初始化列表，Key值用作EXCEL中的标题栏
        bw_data = {'rxkB/s': [], 'txkB/s': [], 'rxMb/s': [], 'txMb/s': []}  # 分类存储服务器sar查询到的信息

        # 通过文件第一行内容定位待提取列索引
        float_data_index = -1
        float_string_index = -1
        rxkB_index = 0
        parts = file.readline().strip().split()
        for part in parts:
            float_string_index += 1
            try:
                float_part = float(part)
                float_data_index += 1
                if float_data_index == 2:
                    rxkB_index = float_string_index
                    break
            except ValueError:
                continue

        for line in file:
            bw_date_str = line.strip().split()
            if bw_date_str and not line.startswith('Average') and not line.startswith('平均'):
                print(bw_date_str)
                rxkB = float(bw_date_str[rxkB_index])
                txkB = float(bw_date_str[rxkB_index + 1])
                rxMb = rxkB * 8 / 1024
                txMb = txkB * 8 / 1024
                bw_data['rxkB/s'].append(rxkB)
                bw_data['txkB/s'].append(txkB)
                bw_data['rxMb/s'].append(rxMb)
                bw_data['txMb/s'].append(txMb)
                logger.debug(f'接收带宽为：{rxMb}, 发送带宽为：{txMb}')
    logger.info(f'完成处理sar数据文件：{sar_file_path}')

    if isAvg:
        # 计算平均值
        rxkB_data = bw_data['rxkB/s']
        txkB_data = bw_data['txkB/s']
        rxMb_data = bw_data['rxMb/s']
        txMb_data = bw_data['txMb/s']

        avg_rxkB = calculate_average(rxkB_data, startIndex, endIndex)
        avg_txkB = calculate_average(txkB_data, startIndex, endIndex)
        avg_rxMb = calculate_average(rxMb_data, startIndex, endIndex)
        avg_txMb = calculate_average(txMb_data, startIndex, endIndex)

        # 将平均值插入到数据顶部
        bw_data['rxkB/s'].insert(0, avg_rxkB)  
        bw_data['txkB/s'].insert(0, avg_txkB)
        bw_data['rxMb/s'].insert(0, avg_rxMb)
        bw_data['txMb/s'].insert(0, avg_txMb)

    # 将处理后的数据写入全局变量
    set_global('bw_data', bw_data)


# 从sar文件中提取数据，该方法适用于未过滤网卡或网口的数据，如：nohup sar -n DEV 1 >> network205SM4.txt &
def extract_sar_datas():
    with open(sar_file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    bw_datas = {}
    n = 0
    for line in lines:
        # isNetworkData = re.search(r'.*时.*分.*秒.*', line)
        # isNetworkData = re.search(r'.*:.*:.*', line)
        isNetworkData = re.search(r'.*:.*:.*|.*时.*分.*秒.*', line)
        if isNetworkData:
            if n <= 1:
                if 'IFACE' in line:
                    n += 1
                else:
                    parts = line.strip().split()
                    if 'AM' in parts or 'PM' in parts:
                        interface = parts[2]
                        logger.info(f'interface is {interface} index is 2')
                    else:
                        interface = parts[1]
                        logger.info(f'interface is {interface} index is 1')
                    bw_datas[interface] = {'rxkB/s': [], 'txkB/s': [], 'rxMb/s': [], 'txMb/s': []}
            elif 'IFACE' not in line:
                parts = line.strip().split()
                if len(parts) >= 6:
                    if 'AM' in parts or 'PM' in parts:
                        current_interface = parts[2]
                        rxkB_s = float(parts[5])
                        txkB_s = float(parts[6])
                    else:
                        current_interface = parts[1]
                        rxkB_s = float(parts[4])
                        txkB_s = float(parts[5])
                    rxMb_s = rxkB_s * 8 / 1024
                    txMb_s = txkB_s * 8 / 1024
                    bw_datas[current_interface]['rxkB/s'].append(rxkB_s)
                    bw_datas[current_interface]['txkB/s'].append(txkB_s)
                    bw_datas[current_interface]['rxMb/s'].append(rxMb_s)
                    bw_datas[current_interface]['txMb/s'].append(txMb_s)
    logger.info(f'完成处理sar数据文件：{sar_file_path}')

    if isAvg:
        # 计算平均值
        for interface in bw_datas.keys():
            rxkB_data = bw_datas[interface]['rxkB/s']
            txkB_data = bw_datas[interface]['txkB/s']
            rxMb_data = bw_datas[interface]['rxMb/s']
            txMb_data = bw_datas[interface]['txMb/s']

            avg_rxkB = calculate_average(rxkB_data, startIndex, endIndex)
            avg_txkB = calculate_average(txkB_data, startIndex, endIndex)
            avg_rxMb = calculate_average(rxMb_data, startIndex, endIndex)
            avg_txMb = calculate_average(txMb_data, startIndex, endIndex)

            # 将平均值插入到数据顶部
            bw_datas[interface]['rxkB/s'].insert(0, avg_rxkB)
            bw_datas[interface]['txkB/s'].insert(0, avg_txkB)
            bw_datas[interface]['rxMb/s'].insert(0, avg_rxMb)
            bw_datas[interface]['txMb/s'].insert(0, avg_txMb)

    # 将处理后的数据写入全局变量
    set_global('bw_datas', bw_datas)
    # print(bw_datas)

    # 将每个接口的数据转换为 DataFrame 并存储
    for interface in bw_datas.keys():
        df = pd.DataFrame(bw_datas[interface])
        set_global(f'{interface}_bw_data_df', df)

def create_folder(folder_name):
    current_directory = os.getcwd()
    new_folder_path = os.path.join(current_directory, folder_name)
    os.makedirs(new_folder_path, exist_ok=True)
    logger.info(f'处理文件存储目录为：{new_folder_path}')

    set_global('new_folder_path', new_folder_path)


def save_top_data():
    logger.info('开始存储top数据excel文件')

    # 从全局变量读取数据
    cpu_data = get_global('cpu_data')
    mem_data = get_global('mem_data')
    pid_data = get_global('pid_data')
    summary_pid_cpu_data = get_global('summary_pid_cpu_data')
    summary_pid_mem_data = get_global('summary_pid_mem_data')

    # print(f'cpu data:\n {cpu_data}')
    # print(f'mem date:\n {mem_data}')
    # print(f'pid data:\n {pid_data}')
    # print(f'sum pid cpu data:\n {summary_pid_cpu_data}')
    # print(f'sum pid mem data:\n {summary_pid_mem_data}')

    file_name = os.path.split(top_file_path)[1].replace('txt', 'xlsx')
    new_file_path = os.path.join(get_global('new_folder_path'), file_name)
    set_global('top_xlsx_file_path', new_file_path)

    # 将数据写入Excel文件
    with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
        # 把dict格式数据转换为pandas DataFrames
        cpu_df = pd.DataFrame(cpu_data)  # 直接使用字典
        mem_df = pd.DataFrame(mem_data)  # 直接使用字典
        set_global('cpu_df', cpu_df)
        set_global('mem_df', mem_df)
        
        # 把CPU和MEM数据写入工作表
        cpu_df.to_excel(writer, sheet_name='CPU', index=True, index_label='NO.')
        mem_df.to_excel(writer, sheet_name='MEM', index=True, index_label='NO.')

        # 把dict格式数据转换为pandas DataFrame
        summary_cpu_df = pd.DataFrame(summary_pid_cpu_data)
        set_global('summary_cpu_df', summary_cpu_df)

        # 把dict格式数据转换为pandas DataFrame
        summary_mem_df = pd.DataFrame(summary_pid_mem_data)
        set_global('summary_mem_df', summary_mem_df)

        # 将汇总数据写入独立的工作表
        summary_cpu_df.to_excel(writer, sheet_name='Summary_CPU', index=True, index_label='NO.')
        summary_mem_df.to_excel(writer, sheet_name='Summary_MEM', index=True, index_label='NO.')

        # 为每个PID创建一个工作表并写入数据
        for pid, data in pid_data.items():
            pid_df = pd.DataFrame(data)
            pid_df.to_excel(writer, sheet_name=pid, index=True, index_label='NO.')

    logger.info('完成存储top数据excel文件')


def save_sar_data():
    logger.info('开始存储sar数据excel文件')

    # 从全局变量读取数据
    bw_data = get_global('bw_data')

    file_name = os.path.split(sar_file_path)[1].replace('txt', 'xlsx')
    new_file_path = os.path.join(get_global('new_folder_path'), file_name)
    set_global('sar_xlsx_file_path', new_file_path)

    # 将数据写入Excel文件
    with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
        # 把dict格式数据转换为pandas DataFrame
        bw_data_df = pd.DataFrame(bw_data)
        set_global('bw_data_df', bw_data_df)

        # 将带宽数据写入独立的工作表
        bw_data_df.to_excel(writer, sheet_name='NetWork', index=True, index_label='NO.')

    logger.info('完成存储sar数据excel文件')


def save_sar_datas():
    logger.info('开始存储sar数据excel文件')

    # 从全局变量读取数据
    bw_datas = get_global('bw_datas')
    if interfaces:
        logger.info(f'提取sar数据：{interfaces}')
        bw_datas = {iface: bw_datas[iface] for iface in interfaces}

    file_name = os.path.split(sar_file_path)[1].replace('txt', 'xlsx')
    new_file_path = os.path.join(get_global('new_folder_path'), file_name)
    set_global('sar_xlsx_file_path', new_file_path)

    workbook = Workbook()
    for interface, values in bw_datas.items():
        # 创建每个 interface 的 sheet
        ws_interface = workbook.create_sheet(title=interface)
        ws_interface.append(['rxkB/s', 'txkB/s', 'rxMb/s', 'txMb/s'])

        for rxkB, txkB, rxMb, txMb in zip(values['rxkB/s'], values['txkB/s'], values['rxMb/s'], values['txMb/s']):
            ws_interface.append([rxkB, txkB, rxMb, txMb])

    # 创建汇总的 sheet
    ws_rx_summary = workbook.create_sheet(title='rxMb_Summary')
    ws_tx_summary = workbook.create_sheet(title='txMb_Summary')

    # 初始化汇总表头
    ws_rx_summary.append([interface for interface in bw_datas.keys()])
    ws_tx_summary.append([interface for interface in bw_datas.keys()])

    # 创建一个字典来存储每个 interface 的 rxMb 和 txMb 数据
    rx_summary_data = {interface: values['rxMb/s'] for interface, values in bw_datas.items()}
    tx_summary_data = {interface: values['txMb/s'] for interface, values in bw_datas.items()}

    # 获取最长的时间长度
    max_len = max(len(values['rxMb/s']) for values in bw_datas.values())

    # 填充汇总数据，按时间顺序
    for i in range(max_len):
        rx_row = [rx_summary_data[interface][i] if i < len(rx_summary_data[interface]) else None for interface in
                  bw_datas.keys()]
        tx_row = [tx_summary_data[interface][i] if i < len(tx_summary_data[interface]) else None for interface in
                  bw_datas.keys()]
        ws_rx_summary.append(rx_row)
        ws_tx_summary.append(tx_row)

    # 删除默认创建的空白 sheet
    workbook.remove(workbook['Sheet'])

    # 保存工作簿到文件
    workbook.save(new_file_path)
    logger.info('完成存储sar数据excel文件')


# 给指定的工作表绘制边框，暂不用
def add_border(writer, sheet_names):
    # 创建边框样式
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    # 获取工作簿对象
    workbook = writer.book
    # 分别获取工作表对象
    for sheet in sheet_names:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border


def figure_chart(chart_title, data_frame_key, columns=None, is_grid=False, xlabel='', ylabel='Usage (%)'):
    logger.debug(f'开始绘图: {chart_title}')

    # 从全局变量获取 DataFrame
    data_frame = get_global(data_frame_key)
    
    if data_frame is None:
        logger.error(f'未找到数据框: {data_frame_key}')
        return

    # 绘制折线图
    plt.figure(figsize=(20, 10))

    # 未指定绘图的数据列，则为整个工作表
    if not columns:
        columns = data_frame.columns

    # 指定每条折线的x,y轴数据，并求取y轴数据最大值
    column_max = []
    for column in columns:
        # 排除第一行（平均值行）
        valid_data = data_frame[column][1:]  # 只取有效数据
        plt.plot(data_frame.index[1:], valid_data, marker=None, label=column)  # x轴也排除第一行
        column_max.append(valid_data.max())
    data_frame_max = max(column_max)

    plt.title(chart_title)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    # 设置x,y轴的显示范围
    y_max = data_frame_max * 1.2
    plt.ylim(0, y_max)
    plt.grid(is_grid)

    # 添加图例，并根据图例项的数量自动换行
    lines_num = len(columns)
    if lines_num > 1:
        legend_items_num = lines_num
        ncol = (legend_items_num + 1) // 2 if legend_items_num > 4 else 4
        plt.legend(loc='upper center', bbox_to_anchor=(0.5, -0.10), ncol=ncol, frameon=False)
        # 调整图表的边缘空间，为图例留出空间
        plt.subplots_adjust(bottom=0.20)

    # 保存折线图为PNG文件
    file_name = f'{chart_title}.png'
    file_path = os.path.join(get_global('new_folder_path'), file_name)
    plt.savefig(file_path)
    logger.debug(f'保存绘图: {chart_title}')

    # 显示折线图
    # plt.show()


def figure_top_chart():
    logger.info('绘制top数据图')
    figure_chart('Server CPU Usage Over Time', 'cpu_df', columns=['used'], is_grid=True)
    figure_chart('Server MEM Usage Over Time', 'mem_df', columns=['used rate'], is_grid=True)
    if pid_list or command_list:
        figure_chart('Process CPU Usage Over Time', 'summary_cpu_df', is_grid=True)
        figure_chart('Process MEM Usage Over Time', 'summary_mem_df', is_grid=True)


def figure_sar_chart():
    logger.info('绘制sar数据图')
    figure_chart('Bandwidth Over Time', 'bw_data_df', columns=['rxMb/s', 'txMb/s'], ylabel='BW(Mb/s)')


def figure_sar_chart_all():
    logger.info('绘制sarAll数据图')
    for interface in interfaces:
        figure_chart(f'{interface} Bandwidth Over Time', f'{interface}_bw_data_df', columns=['rxMb/s', 'txMb/s'], ylabel='BW(Mb/s)')


def process_top_file(pid_list=None, command_list=None):
    try:
        extract_top_data(pid_list, command_list)
        save_top_data()
        figure_top_chart()
    except FileNotFoundError as e:
        logger.error(f'处理top文件时出错: {e}')
        print(f'错误: {e}')


def process_sar_file():
    try:
        extract_sar_data()
        save_sar_data()
        figure_sar_chart()
    except FileNotFoundError as e:
        logger.error(f'处理sar文件时出错: {e}')
        print(f'错误: {e}')


def process_sar_file_all():
    extract_sar_datas()
    save_sar_datas()
    figure_sar_chart_all()


def calculate_average(data, startIndex=None, endIndex=None):
    if startIndex is not None and endIndex is not None:
        # 计算指定索引范围内的平均值
        selected_data = data[startIndex:endIndex]
    elif startIndex is not None:
        # 计算从startIndex到最后的平均值
        selected_data = data[startIndex:]
    elif endIndex is not None:
        # 计算从开始到endIndex的平均值
        selected_data = data[:endIndex]
    else:
        # 计算中间60%的数据
        total_length = len(data)
        start_index = int(total_length * 0.2)  # 20%开始
        end_index = int(total_length * 0.8)    # 80%结束
        selected_data = data[start_index:end_index]

    return round(sum(selected_data) / len(selected_data), 2) if selected_data else 0


if __name__ == '__main__':
    # 待处理的top文件
    top_file_path = r'top.txt'
    # top文件中待处理的Process进程信息
    # 根据PID过滤
    pid_list = [
        7260,7061,7118
    ]
    # 根据Comand过滤，此时pid_list为None或[]
    command_list = ['RLS-MpTas+']
    # 是否计算平均值：
    isAvg = True
    startIndex = None
    endIndex = None

    # 待处理的sar文件
    sar_file_path = r'sar.txt'
    interfaces = ['lo']

    # 新建文件夹保存处理后的数据
    current_time = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    create_folder(current_time)

    # 处理top文件
    process_top_file(pid_list, command_list=command_list)
    # 处理sar文件
    # process_sar_file()           # 按照网卡过滤存储的sar数据用该方法处理，如'sar -n DEV 1 | grep 'bond0' >> sar.txt &'
    process_sar_file_all()       # 全量sar数据用该方法处理，如'sar -n DEV 1 >> sarAll.txt &'
